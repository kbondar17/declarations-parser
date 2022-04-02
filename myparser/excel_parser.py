import re
import statistics
from pathlib import Path
from typing import Union

import pandas as pd
import xlrd
from openpyxl import load_workbook

from myparser.my_logger import get_logger
logger = get_logger(__name__)

class ExcelParser:

    def __init__(self):
        self.cols_we_need = ['department', 'position',
                             'name', 'salary', 'sheet_name', 'documentfile_id']

    @staticmethod
    def find_ok_cols(cols: tuple) -> Union[dict[int, str], None]:
        # отдает словарь {номер_колонки:новое_имя} 
        cols = list(map(str, cols))[:-1]
        cols = list(map(str.lower, cols))
        if len(set(cols)) == 1:
            return

        ok_cols = 0
        result = {}  
        for i, col in enumerate(cols):
            if len(result) > 5 or i > 20:
                return result

            if re.search(pattern='(фамилия|имя|фио|ф\.и\.о\.|ф\.и\.о|отчество)', string=col) and 'name' not in result.values():
                result[i] = 'name'
                ok_cols += 1

            elif re.search(pattern='(рублей|руб|cреднемесячная|зарпл.|плат[ы, е, а]|заработн[ой, ая] плат[а, ы]|cреднемесячн[ая, ой]|зарплат[а, ной, ы])', string=col):
                result[i] = "salary"
                ok_cols += 1

            elif re.search(pattern='(должност[ь, и, ей])', string=col) and 'position' not in result.values():
                result[i] = 'position'
                ok_cols += 1

            elif re.search(pattern='(предприяти[е,я]|учреждени[е,я]|юридическ|организаци|наименование [оу, мо])', string=col) and 'department' not in result.values():
                result[i] = 'department'
                ok_cols += 1

        if len(result) >= 2: # берем ряд, где два и более значения походят на заголовки
            return result
        return False

    def rename_cols(self, df: pd.DataFrame, new_cols: dict[int, str]):
        for k, v in new_cols.items():
            df.rename(columns={df.columns[k]: v}, inplace=True)
        return df
        

    @staticmethod      
    def find_department_in_table(df: pd.DataFrame) -> list[dict[int, str]]:
        """
        Ищем названия учреждения в теле таблицы.
            ФИО       Должость        З/П
                   ФГУП "Ромашка"
            Ваня      Директор       100 тыс.
        отдает [{индекс после которого идет учреждение : учреждение}] 
        """
        departments_n_indexes = []
        for row in df.itertuples():
            index = row[0]
            row = list(row)[1:-1]
            row = [e for e in row if len(str(e)) > 4]
            # если ряд состоит из одинаковых достаточно длинныхзначений,
            # предполагаем, что это учреждение
            if len(set(row)) < 2: 
                if not all([type(e) in [int, float] for e in row]):
                    if statistics.mean([len(e) for e in row]) > 4:
                        departments_n_indexes.append(
                            {'index': index, 'dep': row[0]})

        return departments_n_indexes


    @staticmethod
    def add_department_info_to_df(df: pd.DataFrame, dep_info: list[dict[int, str]]) -> pd.DataFrame:
        df['department'] = None
        for data in dep_info:
            df.at[data['index'], 'department'] = data['dep']
        df = df.dropna(axis=1, how='all').fillna(method='ffill', axis=0)
        return df

    def detect_headers(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Ищет заголовки в теле таблицы. 
        """

        for row in df.itertuples():
            #идем по рядам, ищем заголовки
            # если нашли ФИО, ЗП, учреждение и должность - отдаем df 
            index = row[0]
            new_cols = self.find_ok_cols(row[1:])
            if new_cols:
                logger.debug('Нашли заголовки таблицы --- %s', new_cols)
                break

        if not new_cols:
            raise ValueError('Не нашли заголовки таблицы')
            
        if 'department' not in new_cols.values():
            # если нет колонки с учреждениями - идем искать в таблице
            dep_info = self.find_department_in_table(df)

            if dep_info:
                logger.debug('Нашли учреждения')
                df = self.add_department_info_to_df(df, dep_info)
            else:
                # TODO: Переписать эту дичь
                # если не нашли учреждения в таблице, берем текст перед таблицей
                merged_cells = ' '.join([e for e in set(list(map(str, sum(list(list(e) for e in df.iloc[:index, :].values), [
                ])))) if 'nan' not in e and 'Unnamed' not in e and not e.isalnum() and e not in self.cols_we_need])
                cols = ' '.join([e for e in set(list(map(str, list(df.iloc[:index, :].columns))))
                                if 'nan' not in e and 'Unnamed' not in e and not e.isalnum() and e not in self.cols_we_need])
                df['department'] = cols + ' ' + merged_cells

        df = df.iloc[index+1:, :]
        df = self.rename_cols(df, new_cols)

        return df

    def open_xlsx(self, filename: Path) -> Union[list[pd.DataFrame], None]:
        wb = load_workbook(filename)
        sheets = wb.sheetnames
        dfs = []
        for sheet in sheets:
            df = pd.read_excel(filename, sheet_name=sheet, header=None)
            if not df.empty:
                df['sheet_name'] = sheet
                dfs.append(df)

        if not dfs:
            raise ValueError(f'Не нашли таблиц в {filename}')
        return dfs

    def open_xls(self, filename: Path) -> list[pd.DataFrame]:
        wb = xlrd.open_workbook(filename, ignore_workbook_corruption=True)
        sheets = wb.sheet_names()
        dfs = []

        for sheet in sheets:
            df = pd.read_excel(filename, sheet_name=sheet, header=None)
            if not df.empty:
                df['sheet_name'] = sheet
                dfs.append(df)

        if not dfs:
            raise ValueError(f'Не нашли таблиц в {filename}')
        return dfs

    def parse_file(self, filename: Path) -> list[pd.DataFrame]:
        
        if Path(filename).suffix == '.xlsx':
            dfs = self.open_xlsx(filename)
        elif Path(filename).suffix == '.xls':
            dfs = self.open_xls(filename)
        else:
            raise ValueError('Расширение файла должно быть .xlsx или .xls. file --- %s', filename)

        parsed_dfs = []

        for df in dfs:
            df = df.dropna(axis=1, how='all').fillna(method='ffill', axis=1)
            df = df.fillna(method='ffill', axis=0)
            df = df.fillna(method='ffill', axis=1)
            df = self.detect_headers(df)
            df = df[[c for c in df.columns if c in self.cols_we_need]]
            
            df['documentfile_id'] = filename.name.split('_')[0]
            parsed_dfs.append(df)            

        try:
            return [pd.concat(parsed_dfs).reset_index(drop=True)]
        except:
            return parsed_dfs
